import glob
import os
import shutil
import tempfile
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
import questionary


class ExcelProcessor:
    SKIP_COLORS = {
        "FF00FF00", "0000FF00", "00FF00", "#00FF00",
        "FFFF0000", "00FF0000", "FF0000", "#FF0000",
    }
    GREEN_TARGET = (0, 255, 0)
    RED_TARGET = (255, 0, 0)
    COLOR_TOLERANCE = 110

    def __init__(self, boms_dir: str, ui: Any = None):
        self.boms_dir = boms_dir
        self.ui = ui

    def discover_files(self) -> List[str]:
        search_dir = os.path.join(os.getcwd(), self.boms_dir)
        if not os.path.isdir(search_dir):
            os.makedirs(search_dir, exist_ok=True)
            return []
        return sorted(glob.glob(os.path.join(search_dir, "*.xlsx")))

    def _normalize_color(self, value: Any) -> Optional[str]:
        if not value:
            return None
        raw = str(value).strip()
        if raw.startswith("#"):
            raw = raw[1:]
        if raw.lower().startswith("0x"):
            raw = raw[2:]
        raw = raw.upper()
        if raw == "00000000":
            return None
        if len(raw) == 6:
            return raw
        if len(raw) == 8:
            return raw[2:]
        return None

    def _hex_to_rgb(self, hex_value: str) -> Optional[Tuple[int, int, int]]:
        if not hex_value or len(hex_value) != 6:
            return None
        try:
            return tuple(int(hex_value[i:i+2], 16) for i in (0, 2, 4))
        except ValueError:
            return None

    def _is_close_color(self, rgb: Tuple[int, int, int], target: Tuple[int, int, int]) -> bool:
        distance = sum((c - t) ** 2 for c, t in zip(rgb, target))
        return distance <= self.COLOR_TOLERANCE ** 2

    def _match_skip_color(self, raw: Any) -> Optional[str]:
        normalized = self._normalize_color(raw)
        if normalized in self.SKIP_COLORS:
            if normalized.endswith("00FF00"):
                return "green (done)"
            if normalized.endswith("FF0000"):
                return "red (skip)"

        rgb = self._hex_to_rgb(normalized)
        if rgb is None:
            return None

        if self._is_close_color(rgb, self.GREEN_TARGET):
            return "green (done)"
        if self._is_close_color(rgb, self.RED_TARGET):
            return "red (skip)"
        return None

    def get_cell_color(self, sheet, row: int, col: int = 1) -> Optional[str]:
        try:
            fill = sheet.cell(row=row, column=col).fill
            if not fill or not fill.patternType:
                return None
            color = fill.start_color or fill.fgColor
            if not color:
                return None
            
            raw = None
            if hasattr(color, "rgb") and color.rgb:
                raw = color.rgb
            elif hasattr(color, "index") and color.index:
                raw = color.index
            return self._normalize_color(raw)
        except Exception:
            # Return None if the color value can't be normalized
            return None

    def should_skip_row_color(self, sheet, row: int) -> Optional[str]:
        for col in range(1, min(sheet.max_column, 5) + 1):
            color = self.get_cell_color(sheet, row, col)
            skip = self._match_skip_color(color)
            if skip:
                return skip
        return None

    def _get_wb_for_writing(self, filepath: str):
        # Always reload without data_only=True to preserve formulas
        return openpyxl.load_workbook(filepath)

    def update_cell(self, filepath: str, row: int, col_name: str, value: Any):
        wb = self._get_wb_for_writing(filepath)
        sheet = wb.active
        # Find column index from headers
        headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
        col_idx = headers.index(col_name) + 1
        sheet.cell(row=row, column=col_idx, value=value)
        wb.save(filepath)
        wb.close()

    def mark_row_status(self, filepath: str, row: int, status: str):
        wb = self._get_wb_for_writing(filepath)
        sheet = wb.active
        fill = openpyxl.styles.PatternFill(start_color="00FF00" if status == "OK" else "FF0000", end_color="00FF00" if status == "OK" else "FF0000", fill_type="solid")
        for col in range(1, min(sheet.max_column, 5) + 1):
            sheet.cell(row=row, column=col).fill = fill
        wb.save(filepath)
        wb.close()

    def process_file(self, filepath: str, run_system: str = "ALL", matcher: Any = None, auto_confirm: bool = False) -> tuple[list[dict[str, Any]], dict[str, int]]:
        target_systems = [run_system.upper()] if run_system and run_system != "ALL" else []
            
        stats = {
            "total_excel_rows": 0,
            "empty_rows": 0,
            "example_rows": 0,
            "system_mismatch": 0,
            "skipped_by_color": 0,
            "valid_parts": 0
        }
        
        # Bypassing file lock if needed
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            df = pd.read_excel(filepath)
        except PermissionError:
            if self.ui:
                self.ui.log(f"File '{os.path.basename(filepath)}' is locked. Attempting to read from a temporary copy...", "WARN")
            
            # On Windows, we must close the temp file before copying into it
            fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            
            try:
                shutil.copy2(filepath, temp_path)
                wb = openpyxl.load_workbook(temp_path, data_only=True)
                df = pd.read_excel(temp_path)
            finally:
                if os.path.exists(temp_path):
                    try:
                        os.remove(temp_path)
                    except:
                        pass

        sheet = wb.active
        stats["total_excel_rows"] = len(df)
        
        raw_cols = [str(c).strip().lower() for c in df.columns]
        
        def find_col(aliases):
            for a in aliases:
                if a in raw_cols:
                    return raw_cols.index(a)
            return None

        col_map = {
            "system": find_col(["system", "sys"]),
            "assembly": find_col(["assembly", "asm", "assy"]),
            "part": find_col(["part", "part name", "designation"]),
            "quantity": find_col(["part_quantity", "quantity", "qty", "amount"]),
            "makebuy": find_col(["make o. buy", "m/b", "makebuy"]),
            "comments": find_col(["part_comments", "comments", "notes", "comment"])
        }

        if col_map["system"] is None or col_map["part"] is None:
            raise ValueError(f"Could not identify required 'system' or 'part' columns in {filepath}")

        # DEBUG
        if self.ui:
            self.ui.log(f"DEBUG: raw_cols: {raw_cols}")
            self.ui.log(f"DEBUG: col_map: {col_map}")

        filtered = []
        for idx, row in df.iterrows():
            excel_row = idx + 2
            sys_val = str(row.iloc[col_map["system"]] if col_map["system"] is not None else "").strip().upper()
            
            # Normalize system code (e.g., "C&B" -> "FR")
            if matcher:
                sys_val = matcher.normalize_system_code(sys_val)
                
            part_val = str(row.iloc[col_map["part"]] if col_map["part"] is not None else "").strip()
            
            if not sys_val or sys_val == "NAN" or not part_val or part_val == "NAN":
                stats["empty_rows"] += 1
                continue

            # Check for "BIS HIER NUR BEISPIEL" marker (means everything before was examples)
            row_content = " ".join(str(val) for val in row.values).upper()
            if "BIS HIER" in row_content and "BEISPIEL" in row_content:
                # Add everything processed so far to examples and reset
                stats["example_rows"] += stats["valid_parts"] + stats["system_mismatch"] + stats["skipped_by_color"] + 1
                stats["valid_parts"] = 0
                stats["system_mismatch"] = 0
                stats["skipped_by_color"] = 0
                filtered = []
                continue

            if any(x in sys_val for x in ["BEISPIEL", "EXAMPLE"]) or \
               any(x in part_val.upper() for x in ["BEISPIEL", "EXAMPLE"]):
                stats["example_rows"] += 1
                continue

            if target_systems and sys_val not in target_systems:
                stats["system_mismatch"] += 1
                continue

            if self.should_skip_row_color(sheet, excel_row):
                stats["skipped_by_color"] += 1
                continue

            asm_val = str(row.iloc[col_map["assembly"]] if col_map["assembly"] is not None else "").strip()
            qty_val = str(row.iloc[col_map["quantity"]] if col_map["quantity"] is not None else "").strip()
            mb_val = str(row.iloc[col_map["makebuy"]] if col_map["makebuy"] is not None else "m").strip().lower()[:1] or "m"
            comm_val = str(row.iloc[col_map["comments"]] if col_map["comments"] is not None else "").strip().replace("nan", "")

            # --- Validation: Part Name (Max 25) ---
            if len(part_val) > 25:
                res = self._handle_long_field(filepath, part_val, 25, "Part Name", excel_row, auto_confirm)
                if res is None: # Skip
                    continue
                part_val = res

            # --- Validation: Comment (Max 40) ---
            if len(comm_val) > 40:
                res = self._handle_long_field(filepath, comm_val, 40, "Comment", excel_row, auto_confirm)
                if res is None: # Skip
                    continue
                comm_val = res

            filtered.append({
                "row": excel_row,
                "system": sys_val,
                "assembly": asm_val,
                "part": part_val,
                "makebuy": mb_val,
                "quantity": qty_val,
                "comments": comm_val,
            })
            stats["valid_parts"] += 1

        return filtered, stats

    def _handle_long_field(self, filepath: str, value: str, limit: int, field_name: str, row: int, auto_confirm: bool) -> str | None:
        """Helper to handle long text fields interactively or automatically."""
        if auto_confirm:
            if self.ui:
                self.ui.log(f"Row {row}: {field_name} too long ({len(value)} > {limit}). Auto-skipping.", "WARN")
            with open("bom_log.txt", "a") as log:
                log.write(f"Row {row}: Action=Auto-Skip ({field_name}), Original='{value}'\n")
            return None

        from rich.console import Console
        console = Console()

        while True:
            console.print(f"\n[bold yellow][!] {field_name} too long at row {row}[/]")
            color = "red" if len(value) > limit else "green"
            console.print(f"Current: [italic]{value}[/] ([{color}]{len(value)}[white]/{limit} chars)")
            
            choice = questionary.select(
                f"How would you like to handle this {field_name.lower()}?",
                choices=[
                    "Edit (with pre-fill)",
                    f"Truncate to {limit} chars",
                    "Skip row - do NOT upload"
                ]
            ).ask()

            col_name = "part" if field_name == "Part Name" else "part_comments"

            if choice == "Edit (with pre-fill)":
                def validate_len(text):
                    if len(text) <= limit:
                        return True
                    return f"Too long! ({len(text)}/{limit} chars)"

                edited = questionary.text(
                    f"Enter new {field_name.lower()} (max {limit}):",
                    default=value,
                    validate=validate_len,
                    instruction=" (use arrow keys to go back or ESC to cancel editing)"
                ).ask()
                
                if edited is None: # User pressed ESC or similar
                    continue # Loop back to choice menu
                
                # Check if they want to go back
                back_confirm = questionary.confirm("Keep this edit?", default=True).ask()
                if not back_confirm:
                    continue
                
                value = edited
                self.update_cell(filepath, row, col_name, value)
                with open("bom_log.txt", "a") as log:
                    log.write(f"Row {row}: Action=Edit ({field_name}), Final='{value}'\n")
                return value

            elif choice == f"Truncate to {limit} chars":
                value = value[:limit]
                self.update_cell(filepath, row, col_name, value)
                with open("bom_log.txt", "a") as log:
                    log.write(f"Row {row}: Action=Truncate ({field_name}), Final='{value}'\n")
                return value

            elif choice == "Skip row - do NOT upload":
                with open("bom_log.txt", "a") as log:
                    log.write(f"Row {row}: Action=Skip ({field_name}), Original='{value}'\n")
                return None
