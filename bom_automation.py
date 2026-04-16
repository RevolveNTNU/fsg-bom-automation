"""
FSG CCBOM Automation Tool
=========================

Automates the upload of Costed Carbonized Bill of Material (CCBOM) data
from Excel files to the Formula Student Germany (FSG) website.

Developed by ELBFLORACE e.V. — Dresden, Germany.
Open-sourced for all Formula Student teams.

Usage:
    python bom_automation.py

Configuration:
    See .env.example for all available options.
"""

import os
import sys
import glob
import time
import pandas as pd
import openpyxl
import re
import random
import threading
from datetime import datetime
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright

# ──────────────────────────────────────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────────────────────────────────────

load_dotenv()

# FSG Website URLs — TEAM_ID must be explicitly set in .env
TEAM_ID = os.getenv("TEAM_ID", "").strip()
if not TEAM_ID:
    print("ERROR: TEAM_ID not set. Copy .env.example to .env and set TEAM_ID.")
    sys.exit(1)
BASE_URL = "https://www.formulastudent.de"
LOGIN_URL = f"{BASE_URL}/login"
BOM_URL = f"{BASE_URL}/teams/fse/details/bom/tid/{TEAM_ID}"

# Credentials & Behaviour
FSG_USERNAME = os.getenv("FSG_USERNAME")
FSG_PASSWORD = os.getenv("FSG_PASSWORD")
# Safer default for new users: TEST_MODE enabled by default.
TEST_MODE = os.getenv("TEST_MODE", "true").lower() == "true"
# Dry-run option: if true, log actions but do not perform uploads.
DRY_RUN = os.getenv("DRY_RUN", "false").lower() == "true"
TEST_LIMIT = int(os.getenv("TEST_LIMIT", "3"))
DEFAULT_SYSTEM = os.getenv("DEFAULT_SYSTEM", "").strip().upper()
ALLOWED_ASSEMBLIES_RAW = os.getenv("ALLOWED_ASSEMBLIES", "").strip()
ALLOWED_ASSEMBLIES = [a.strip() for a in ALLOWED_ASSEMBLIES_RAW.split(",") if a.strip()]
LOG_FILE = os.getenv("LOG_FILE", "bom_log.txt")
BOMS_DIR = os.getenv("BOMS_DIR", "BOMs")
PART_MAX_LENGTH = 25
COMMENTS_MAX_LENGTH = 40

# Rate Limiting & Robustness
BASE_DELAY = float(os.getenv("BASE_DELAY", "2.0"))
BURST_LIMIT = int(os.getenv("BURST_LIMIT", "10"))
BURST_COOLDOWN = float(os.getenv("BURST_COOLDOWN", "15.0"))
MAX_RETRIES = int(os.getenv("MAX_RETRIES", "2"))
REQUEST_TIMEOUT = float(os.getenv("REQUEST_TIMEOUT", "30.0")) * 1000 # convert to ms

# Row‑colour hex codes used for skip detection (openpyxl format, no leading #)
SKIP_COLORS = {
    "FF00FF00",   # Pure Green  — already uploaded
    "0000FF00",   # Green variant
    "#00ff00",   # Another green variant (some Excel versions)
    "FFFF0000",   # Pure Red    — do not upload
    "00FF0000",   # Red variant
    "#ff0000",   # Another red variant (some Excel versions)
}

# ──────────────────────────────────────────────────────────────────────────────
# FSG System Codes → Full Dropdown Labels
# ──────────────────────────────────────────────────────────────────────────────

SYSTEM_MAP = {
    "AT": "AT - Autonomous System",
    "BR": "BR - Brake System",
    "DT": "DT - Drivetrain",
    "ET": "ET - Engine and Tractive System",
    "FR": "FR - Chassis and Body",
    "LV": "LV - Grounded Low Voltage System",
    "MS": "MS - Miscellaneous Fit and Finish",
    "ST": "ST - Steering System",
    "SU": "SU - Suspension System",
    "WT": "WT - Wheels, Wheel Bearings and Tires",
}

# ──────────────────────────────────────────────────────────────────────────────
# Smart Assembly Remapping
#
# If your Excel uses a different name for an assembly than the FSG dropdown,
# add the mapping here.  Keys must be lowercase.
# ──────────────────────────────────────────────────────────────────────────────

ASSEMBLY_REMAP = {
    # Brake System
    "brake caliper":        "Calipers",
    "brake calipers":       "Calipers",
    "caliper":              "Calipers",
    "reservoire":           "Brake Master Cylinder",
    "reservoir":            "Brake Master Cylinder",
    "resovoir":             "Brake Master Cylinder",
    "fitting screw":        "Fasteners",
    "fastener":             "Fasteners",
    "screws":               "Fasteners",
    "bolts":                "Fasteners",
    "brake disc":           "Brake Discs",
    "brake disk":           "Brake Discs",
    "brake pad":            "Brake Pads",
    "brake line":           "Brake Lines",
    "master cylinder":      "Brake Master Cylinder",
    # Suspension
    "damper":               "Dampers",
    "spring":               "Springs",
    "pushrod":              "Pushrods",
    "rocker":               "Rockers",
    "a-arm":                "A-Arms",
    # Drivetrain
    "chain":                "Chain",
    "sprocket":             "Sprockets",
    "differential":         "Differential",
    "half shaft":           "Half Shafts",
    "halfshaft":            "Half Shafts",
    # Steering
    "steering rack":        "Steering Rack",
    "tie rod":              "Tie Rods",
    "steering wheel":       "Steering Wheel",
    # Wheels & Tires
    "tire":                 "Tires",
    "tyre":                 "Tires",
    "wheel bearing":        "Wheel Bearings",
    "rim":                  "Wheels",
    "wheel":                "Wheels",
}


# ──────────────────────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────────────────────

def smart_delay(seconds, jitter=0.2):
    """Sleep for N seconds with ±jitter percentage."""
    if seconds <= 0:
        return
    actual = seconds * (1 + random.uniform(-jitter, jitter))
    time.sleep(max(0.1, actual))

def prompt_ask(question):
    """
    Run a questionary prompt in a separate thread to avoid 
    asyncio event loop conflicts with Playwright.
    """
    result = [None]
    exception = [None]

    def _target():
        try:
            result[0] = question.ask()
        except Exception as e:
            exception[0] = e

    # Use a daemon thread and join in short intervals so KeyboardInterrupt
    # (Ctrl+C) is handled by the main thread while the prompt is active.
    thread = threading.Thread(target=_target, daemon=True)
    thread.start()
    try:
        while thread.is_alive():
            thread.join(timeout=0.1)
    except KeyboardInterrupt:
        # Let the main thread handle the interrupt; daemon thread won't block exit.
        raise

    if exception[0]:
        raise exception[0]
    return result[0]

def log(message: str, status: str = "INFO") -> None:
    """Log to both console and log file with timestamp."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] [{status}] {message}"
    print(line)
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


# ──────────────────────────────────────────────────────────────────────────────
# Excel Helpers
# ──────────────────────────────────────────────────────────────────────────────

def normalize_color(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).upper()
    if raw == "00000000":
        return None
    if len(raw) == 6:
        raw = "FF" + raw
    if len(raw) == 8:
        return raw
    return raw


def get_cell_color(sheet, row: int, col: int = 1) -> str | None:
    """Return the fill colour of a cell as an uppercase hex string, or None."""
    try:
        fill = sheet.cell(row=row, column=col).fill
        if fill is None or fill.patternType is None:
            return None

        color = fill.start_color or fill.fgColor
        if color is None:
            return None

        # Prefer explicit RGB if available.
        raw = None
        if hasattr(color, "rgb") and color.rgb:
            raw = color.rgb
        elif hasattr(color, "index") and color.index:
            raw = color.index
        elif hasattr(color, "indexed") and color.indexed is not None:
            raw = color.indexed
        elif hasattr(color, "theme") and color.theme is not None:
            raw = color.theme

        return normalize_color(raw)
    except Exception:
        return None


def should_skip_color(sheet, row: int) -> str | None:
    """Check if a row's colour indicates it should be skipped.
    Returns a human-readable reason or None."""
    for col in range(1, min(sheet.max_column, 10) + 1):
        color = get_cell_color(sheet, row, col)
        if not color:
            continue
        if color in SKIP_COLORS:
            if color in {"FF00FF00", "0000FF00"}:
                return "green (already uploaded)"
            if color in {"FFFF0000", "00FF0000"}:
                return "red (do not upload)"
            return f"skipped colour ({color})"
    return None


def init_excel_audit_columns(sheet):
    """Ensure the Excel sheet has audit columns at the end."""
    headers = [cell.value for cell in sheet[1]]
    
    if "Upload Status" not in headers:
        status_col = sheet.max_column + 1
        sheet.cell(row=1, column=status_col, value="Upload Status")
    else:
        status_col = headers.index("Upload Status") + 1
        
    if "Automation Notes" not in headers:
        notes_col = sheet.max_column + 1
        sheet.cell(row=1, column=notes_col, value="Automation Notes")
    else:
        notes_col = headers.index("Automation Notes") + 1
        
    return status_col, notes_col


def perform_data_surgery(filtered_list, wb, filepath):
    """Interactively fix overlong Part Names or Comments."""
    from rich.console import Console
    from rich.panel import Panel
    console = Console()
    
    sheet = wb.active
    edited = False
    
    # Map column names to indices for updating
    headers = [str(cell.value).lower().strip() for cell in sheet[1]]
    part_col = headers.index("part") + 1 if "part" in headers else None
    comments_col = headers.index("part_comments") + 1 if "part_comments" in headers else None

    if part_col is None:
        return # Should not happen based on previous checks

    to_remove = []

    for item in filtered_list:
        part_len = len(item["part"])
        comments_len = len(item["comments"])
        
        issue = ""
        if part_len > PART_MAX_LENGTH:
            issue = f"Part Name too long ({part_len} > {PART_MAX_LENGTH})"
        elif comments_len > COMMENTS_MAX_LENGTH:
            issue = f"Comments too long ({comments_len} > {COMMENTS_MAX_LENGTH})"
            
        if issue:
            console.print("\n")
            console.print(Panel(
                f"[bold red]Row {item['row']}: {issue}[/]\n"
                f"[bold]Part:[/] {item['part']}\n"
                f"[bold]Comments:[/] {item['comments'] or 'N/A'}",
                title="Data Length Violation"
            ))
            
            action = prompt_ask(
                questionary.select(
                    "Choose action:",
                    choices=[
                        questionary.Choice("Auto-Truncate", "truncate"),
                        questionary.Choice("Manual Edit", "manual"),
                        questionary.Choice("Skip this row", "skip")
                    ]
                )
            )
            
            if action == "truncate":
                item["part"] = item["part"][:PART_MAX_LENGTH].strip()
                item["comments"] = (item["comments"] or "")[:COMMENTS_MAX_LENGTH].strip()
                sheet.cell(row=item["row"], column=part_col, value=item["part"])
                if comments_col:
                    sheet.cell(row=item["row"], column=comments_col, value=item["comments"])
                edited = True
                log(f"Row {item['row']}: Auto-truncated", "FIX")
                
            elif action == "manual":
                new_part = prompt_ask(
                    questionary.text(
                        f"New Part Name (max {PART_MAX_LENGTH}):",
                        default=item["part"][:PART_MAX_LENGTH],
                        validate=lambda text: len(text) <= PART_MAX_LENGTH or f"Too long! ({len(text)}/{PART_MAX_LENGTH})"
                    )
                )
                
                new_comments = ""
                if comments_col:
                    new_comments = prompt_ask(
                        questionary.text(
                            f"New Comments (max {COMMENTS_MAX_LENGTH}):",
                            default=(item["comments"] or "")[:COMMENTS_MAX_LENGTH],
                            validate=lambda text: len(text) <= COMMENTS_MAX_LENGTH or f"Too long! ({len(text)}/{COMMENTS_MAX_LENGTH})"
                        )
                    )
                
                item["part"] = new_part
                item["comments"] = new_comments
                sheet.cell(row=item["row"], column=part_col, value=new_part)
                if comments_col:
                    sheet.cell(row=item["row"], column=comments_col, value=new_comments)
                edited = True
                log(f"Row {item['row']}: Manually edited", "FIX")
                
            elif action == "skip":
                to_remove.append(item)
                log(f"Row {item['row']}: Skipped by user during surgery", "SKIP")

    # Remove skipped items
    for item in to_remove:
        filtered_list.remove(item)
        
    if edited:
        wb.save(filepath)
        log(f"Changes saved back to {filepath}")

# ──────────────────────────────────────────────────────────────────────────────
# Fuzzy Dropdown Matching
# ──────────────────────────────────────────────────────────────────────────────

def fuzzy_select(page, selector: str, target: str) -> bool:
    """Select a dropdown option using smart remapping + fuzzy matching.

    Matching priority:
      0. Check ASSEMBLY_REMAP dictionary
      1. Exact match
      2. Case-insensitive match
      3. Substring / contains match
    """
    try:
        # 0 — Remap
        resolved = ASSEMBLY_REMAP.get(target.lower().strip(), target)

        options = page.eval_on_selector(
            selector,
            "el => Array.from(el.options).map(o => o.text)",
        )

        # 1 — Exact
        if resolved in options:
            page.locator(selector).select_option(label=resolved)
            return True

        # 2 — Case-insensitive
        resolved_lower = resolved.lower().strip()
        for opt in options:
            if opt.lower().strip() == resolved_lower:
                page.locator(selector).select_option(label=opt)
                return True

        # 3 — Contains
        for opt in options:
            ol = opt.lower().strip()
            if resolved_lower in ol or ol in resolved_lower:
                page.locator(selector).select_option(label=opt)
                return True

        return False
    except Exception as e:
        log(f"Fuzzy-match error for '{target}': {e}", "WARN")
        return False


def resolve_assembly_label(target: str, options: list[str], allowed: list[str] | None = None) -> str | None:
    """Match an Excel assembly value to one of the site options, optionally restricted by allowed labels."""
    resolved = ASSEMBLY_REMAP.get(target.lower().strip(), target).strip()
    target_lower = resolved.lower()
    # normalized form (remove non-alphanumeric) to match things like 'gear box' vs 'gearbox'
    def _norm(s: str) -> str:
        return re.sub(r"\W+", "", str(s or "").lower())
    target_norm = _norm(resolved)
    if allowed:
        allowed_set = {a.strip().lower() for a in allowed if a.strip()}
        options = [opt for opt in options if opt.strip().lower() in allowed_set]
    if not options:
        return None
    # Exact match first
    for opt in options:
        if opt.strip() == resolved:
            return opt
    for opt in options:
        if opt.strip().lower() == target_lower:
            return opt
    for opt in options:
        if target_lower in opt.strip().lower() or opt.strip().lower() in target_lower:
            return opt
    # Try normalized comparisons (ignore spaces/punctuation)
    for opt in options:
        if _norm(opt) == target_norm:
            return opt
    for opt in options:
        on = _norm(opt)
        if target_norm in on or on in target_norm:
            return opt
    return None


# ──────────────────────────────────────────────────────────────────────────────
# File Selection
# ──────────────────────────────────────────────────────────────────────────────

def discover_excel_files() -> list[str]:
    """Look for .xlsx files inside the BOMS_DIR folder."""
    search_dir = os.path.join(os.getcwd(), BOMS_DIR)
    if not os.path.isdir(search_dir):
        os.makedirs(search_dir, exist_ok=True)
        log(f"Created '{BOMS_DIR}/' directory. Place your Excel files there.", "WARN")
        return []
    files = glob.glob(os.path.join(search_dir, "*.xlsx"))
    return sorted(files)


import questionary

def select_file() -> str:
    """Interactive file picker for Excel BOMs using questionary."""
    files = discover_excel_files()
    if not files:
        from rich.console import Console
        console = Console()
        console.print(f"\n[red]No .xlsx files found in '{BOMS_DIR}/'.[/]")
        sys.exit(1)

    choices = [os.path.basename(f) for f in files]
    selected_name = prompt_ask(
        questionary.select(
            "Select the BOM Excel file:",
            choices=choices
        )
    )
    
    if not selected_name:
        sys.exit(0)
        
    for f in files:
        if os.path.basename(f) == selected_name:
            return f
    return files[0]


from rich.table import Table
from rich.console import Console
from rich.panel import Panel

def show_summary(filtered_count, filename, system, test_mode, dry_run):
    console = Console()
    table = Table(title="Upload Configuration Summary", show_header=False)
    table.add_row("File", filename)
    table.add_row("System", system)
    table.add_row("Parts to Upload", str(filtered_count))
    table.add_row("Test Mode", "[yellow]ENABLED[/]" if test_mode else "[green]DISABLED[/]")
    table.add_row("Dry Run", "[yellow]ENABLED[/]" if dry_run else "[green]DISABLED[/]")
    
    console.print("\n")
    console.print(Panel(table, title="Ready to Upload", expand=False))

# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

def main() -> None:
    from rich.console import Group
    log("=" * 60)
    log("FSG CCBOM Automation — Starting")
    if TEST_MODE:
        log(f"TEST MODE: Only the first {TEST_LIMIT} parts will be processed.", "WARN")
    log(f"Config: TEAM_ID={TEAM_ID} TEST_MODE={TEST_MODE} DRY_RUN={DRY_RUN} BOMS_DIR={BOMS_DIR}")

    # ── 1. File selection ────────────────────────────────────────────────
    filepath = select_file()
    filename = os.path.basename(filepath)
    log(f"Selected file: {filename}")

    wb = openpyxl.load_workbook(filepath, data_only=False) # data_only=False to preserve formulas and allow saving
    sheet = wb.active
    
    # Initialize audit columns
    status_col, notes_col = init_excel_audit_columns(sheet)
    wb.save(filepath)
    
    # Reload for processing with data_only=True if needed, but we'll try to stick to one load
    df = pd.read_excel(filepath)

    # Normalise column names (strip whitespace, lowercase)
    df.columns = [c.strip().lower() for c in df.columns]

    required_cols = {"system", "assembly", "part"}
    missing = required_cols - set(df.columns)
    if missing:
        log(f"Missing required columns: {missing}. "
            f"Available: {list(df.columns)}", "ERROR")
        sys.exit(1)

    # ── 2. System selection ──────────────────────────────────────────────
    unique_systems = [
        str(s).strip().upper()
        for s in df["system"].dropna().unique()
        if str(s).strip().upper() not in ("NAN", "BEISPIEL", "")
    ]

    if DEFAULT_SYSTEM and DEFAULT_SYSTEM in unique_systems:
        run_system = DEFAULT_SYSTEM
        log(f"System filter (from .env): {run_system} — "
            f"{SYSTEM_MAP.get(run_system, run_system)}")
    else:
        system_choices = [
            questionary.Choice(title=f"{s} - {SYSTEM_MAP.get(s, s)}", value=s)
            for s in unique_systems
        ]
        system_choices.insert(0, questionary.Choice(title="ALL - Process everything", value="ALL"))
        
        run_system = prompt_ask(
            questionary.select(
                "Select the system to process:",
                choices=system_choices
            )
        )
        
        if not run_system:
            sys.exit(0)

    # ── 3. Filter rows ───────────────────────────────────────────────────
    filtered = []
    skipped_green = 0
    skipped_red = 0
    skipped_example = 0
    skipped_empty = 0

    for idx, row in df.iterrows():
        excel_row = idx + 2  # pandas 0-indexed + header row

        system = str(row.get("system", "")).strip().upper()
        assembly = str(row.get("assembly", "")).strip()
        part = str(row.get("part", "")).strip()
        quantity = row.get("part_quantity", "")
        makebuy = str(row.get("make o. buy", "")).strip().lower()
        comments = str(row.get("part_comments", "")).strip()

        # System filter
        if run_system != "ALL" and system != run_system:
            continue

        # Empty rows
        if (pd.isna(row.get("system")) and pd.isna(row.get("part"))) or \
           not system or system == "NAN" or not part or part == "NAN":
            skipped_empty += 1
            continue

        # Example rows
        if "BEISPIEL" in system or "BEISPIEL" in part.upper() or \
           "EXAMPLE" in system or "EXAMPLE" in part.upper():
            skipped_example += 1
            continue

        # Colour check (green = uploaded, red = do not upload)
        skip_reason = should_skip_color(sheet, excel_row)
        if skip_reason:
            if "green" in skip_reason:
                skipped_green += 1
            elif "red" in skip_reason:
                skipped_red += 1
            log(f"Row {excel_row}: Skipped — {skip_reason}", "SKIP")
            continue

        # Clean make/buy to single char
        mb = makebuy[0] if makebuy and makebuy[0] in ("m", "b") else "m"

        # Clean comments
        if comments in ("nan", "NaN", ""):
            comments = ""

        # Clean quantity
        qty_str = str(quantity).strip()
        if qty_str in ("nan", "NaN", ""):
            qty_str = ""

        # Field length checks: the website rejects overlong values.
        part_len = len(part)
        comments_len = len(comments)
        if part_len > PART_MAX_LENGTH:
            log(
                f"Row {excel_row}: Part too long ({part_len} > {PART_MAX_LENGTH}) — '{part[:PART_MAX_LENGTH]}...' not uploaded",
                "ERROR",
            )
            continue
        if comments_len > COMMENTS_MAX_LENGTH:
            log(
                f"Row {excel_row}: Comments too long ({comments_len} > {COMMENTS_MAX_LENGTH}) — '{comments[:COMMENTS_MAX_LENGTH]}...' not uploaded",
                "ERROR",
            )
            continue

        filtered.append({
            "row": excel_row,
            "system": system,
            "assembly": assembly,
            "part": part,
            "makebuy": mb,
            "quantity": qty_str,
            "comments": comments,
        })

    # Summary
    log(f"Filtering complete: {len(filtered)} parts to upload "
        f"({skipped_green} green / {skipped_red} red / "
        f"{skipped_example} example / {skipped_empty} empty skipped)")

    if not filtered:
        log("Nothing to upload — exiting.")
        sys.exit(0)

    # ── 3b. Pre-flight data surgery ──────────────────────────────────────
    perform_data_surgery(filtered, wb, filepath)
    
    if not filtered:
        log("Nothing left to upload after data surgery — exiting.")
        sys.exit(0)

    if not (FSG_USERNAME and FSG_PASSWORD):
        from rich.console import Console
        console = Console()
        console.print(Panel(
            "[yellow]FSG_USERNAME and/or FSG_PASSWORD not set.[/]\n"
            "The script will open a browser and you will need to log in manually.",
            title="Manual Login Mode"
        ))
        if not prompt_ask(questionary.confirm("Continue in manual login mode?")):
            log("Aborted by user: credentials missing and manual login declined.", "ERROR")
            sys.exit(1)

    # Final pre-upload confirmation
    show_summary(
        len(filtered), 
        filename, 
        run_system if run_system != "ALL" else "ALL", 
        TEST_MODE, 
        DRY_RUN
    )
    
    if not prompt_ask(questionary.confirm("Proceed with uploading?")):
        log("Aborted by user before upload.", "WARN")
        sys.exit(0)

    # ── 4. Browser automation ────────────────────────────────────────────
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()

        # Login
        if FSG_USERNAME and FSG_PASSWORD:
            log(f"Logging in as '{FSG_USERNAME}'...")
            page.goto(LOGIN_URL)
            page.fill("#tx-felogin-input-username", FSG_USERNAME)
            page.fill("#tx-felogin-input-password", FSG_PASSWORD)
            page.click('input[name="submit"]')
            page.wait_for_load_state("networkidle")

        page.goto(BOM_URL)
        from rich.panel import Panel
        from rich.console import Console
        console = Console()
        console.print(Panel(
            "Verify you are logged in and on the BOM page.\n"
            "Press ENTER to begin fetching assembly options.",
            title="Action Required"
        ))
        input()

        site_options = []
        try:
            page.get_by_text("New", exact=True).click()
            page.wait_for_selector(".DTE_Action_Create", timeout=5000)

            # Read available systems on the site and try to pre-select the
            # system the user chose earlier (run_system).
            try:
                site_systems = page.eval_on_selector(
                    "#DTE_Field_system",
                    "el => Array.from(el.options).map(o => o.text)",
                ) or []
            except Exception:
                site_systems = []

            try:
                if run_system and run_system != "ALL":
                    sys_label = SYSTEM_MAP.get(run_system, run_system)
                    selected = False
                    try:
                        selected = fuzzy_select(page, "#DTE_Field_system", sys_label)
                    except Exception:
                        selected = False

                    if not selected and site_systems:
                        # Offer the user a choice of which site system to use
                        print("\nCould not auto-select system in the dialog. Choose site system to filter assemblies:")
                        for i, s in enumerate(site_systems, start=1):
                            print(f"  {i:2d}. {s}")
                        choice = input("Enter number to select system (or press ENTER to skip): ").strip()
                        try:
                            if choice:
                                n = int(choice)
                                if 1 <= n <= len(site_systems):
                                    sel = site_systems[n - 1]
                                    fuzzy_select(page, "#DTE_Field_system", sel)
                                    page.locator("#DTE_Field_system").dispatch_event("change")
                                    page.wait_for_timeout(300)
                        except Exception:
                            pass
                    else:
                        try:
                            page.locator("#DTE_Field_system").dispatch_event("change")
                            page.wait_for_timeout(300)
                        except Exception:
                            pass
            except Exception as e:
                log(f"Could not pre-select system in dialog: {e}", "WARN")

            site_options = page.eval_on_selector(
                "#DTE_Field_assembly",
                "el => Array.from(el.options).map(o => o.text)",
            ) or []
        except Exception as e:
            log(f"Could not read assembly options from the site: {e}", "WARN")
            site_options = []
        finally:
            try:
                page.keyboard.press("Escape")
                page.wait_for_timeout(200)
            except Exception:
                pass

        # ── 4b. Interactive Assembly Selection (STAYS INSIDE SINGLE PW SESSION) ─────────
        runtime_allowed: list[str] = []
        if site_options:
            if ALLOWED_ASSEMBLIES:
                allowed_lower = {a.strip().lower() for a in ALLOWED_ASSEMBLIES}
                runtime_allowed = [opt for opt in site_options if opt.strip().lower() in allowed_lower]
                if not runtime_allowed:
                    log("ALLOWED_ASSEMBLIES set but none matched available site options; continuing with all options.", "WARN")
                    runtime_allowed = []
            else:
                # Use our robust, sync-safe prompt_ask helper to wrap the checkbox.
                # This explicitly handles the thread-local nature and avoids the
                # runtime error by not relying on the same event loop context.
                from questionary import checkbox

                runtime_allowed = prompt_ask(
                    checkbox(
                        "Select assemblies to upload (Space to toggle, Enter to confirm):",
                        choices=site_options,
                        validate=lambda a: (len(a) > 0) or "Select at least one assembly"
                    )
                )

            # Apply mapping / filtering based on selected assemblies
            new_filtered: list[dict] = []
            skipped_mapping = 0
            if runtime_allowed:
                for item in filtered:
                    resolved = resolve_assembly_label(item["assembly"], site_options, runtime_allowed)
                    if resolved:
                        item["assembly"] = resolved
                        new_filtered.append(item)
                    else:
                        log(f"Row {item['row']}: Assembly '{item['assembly']}' not in selected assemblies — skipping", "WARN")
                        skipped_mapping += 1
            else:
                for item in filtered:
                    resolved = resolve_assembly_label(item["assembly"], site_options, None)
                    if resolved:
                        item["assembly"] = resolved
                        new_filtered.append(item)
                    else:
                        log(f"Row {item['row']}: Assembly '{item['assembly']}' not found in dropdown — skipping", "WARN")
                        skipped_mapping += 1

            filtered = new_filtered
            log(f"Selected assemblies: {runtime_allowed or ['ALL']}")
            log(f"After assembly selection: {len(filtered)} parts to upload ({skipped_mapping} skipped due to mapping).")
            if TEST_MODE and len(filtered) > TEST_LIMIT:
                log(f"Test Mode: limiting {len(filtered)} → {TEST_LIMIT} parts")
                filtered = filtered[:TEST_LIMIT]
            if not filtered:
                log("No rows left to upload after assembly selection — exiting.", "ERROR")
                sys.exit(0)
        else:
            log("No assembly options were found on the BOM page.", "WARN")

        # ── 5. Deduplication ─────────────────────────────────────────────
        log("Fetching existing parts for deduplication...")
        existing: dict[str, dict] = {} # Key -> Full Data
        
        # Wait for the table to appear and rows to load (or the "empty" message)
        try:
            page.wait_for_selector("#bom-table", timeout=10000)
            # Short wait to allow AJAX/rendering to finish
            page.wait_for_timeout(2000)
        except Exception:
            log("Deduplication: Table #bom-table not found on page.", "WARN")

        try:
            data = page.evaluate(
                """() => {
                    const results = [];
                    try {
                        const tableEl = document.querySelector('#bom-table');
                        if (!tableEl) return [];

                        // 1. Try DataTable API first (best for all pages)
                        if (typeof $ !== 'undefined' && $.fn.DataTable && $.fn.dataTable.isDataTable('#bom-table')) {
                            const dt = $('#bom-table').DataTable();
                            const allData = dt.data().toArray();
                            const headers = dt.columns().header().toArray().map(h => h.innerText.toLowerCase().trim());
                            
                            allData.forEach(row => {
                                const obj = {};
                                if (Array.isArray(row)) {
                                    row.forEach((val, i) => {
                                        const h = headers[i];
                                        if (h) obj[h] = val;
                                    });
                                } else {
                                    Object.assign(obj, row);
                                }
                                results.push(obj);
                            });
                        } 
                        
                        // 2. Fallback: Scrape visible HTML rows if DataTable API didn't return data
                        if (results.length === 0) {
                            const headerEls = Array.from(tableEl.querySelectorAll('thead th')).map(th => th.innerText.toLowerCase().trim());
                            const rows = tableEl.querySelectorAll('tbody tr');
                            rows.forEach(tr => {
                                if (tr.classList.contains('empty') || tr.innerText.includes('No data')) return;
                                const obj = {};
                                const cells = tr.querySelectorAll('td');
                                cells.forEach((td, i) => {
                                    const h = headerEls[i];
                                    if (h) obj[h] = td.innerText.trim();
                                });
                                // Capture Row ID if present
                                if (tr.id) obj['DT_RowId'] = tr.id;
                                results.push(obj);
                            });
                        }
                    } catch(e) {
                        return {error: e.message};
                    }
                    return results;
                }"""
            )
            
            if isinstance(data, dict) and "error" in data:
                log(f"Scraping JS Error: {data['error']}", "WARN")
                data = []

            for r in data:
                if isinstance(r, dict):
                    # Normalized field extraction
                    # Site labels: "system", "assembly", "part", "quantity", "make o. buy", "comments"
                    sys_val = str(r.get('system') or '').split(' - ')[0].strip()
                    # Fallback for system if empty (sometimes it's in the row ID: "SU_123")
                    if not sys_val and r.get('DT_RowId'):
                        sys_val = str(r.get('DT_RowId')).split('_')[0].strip()
                        
                    asm_val = str(r.get('assembly') or '').strip()
                    part_val = str(r.get('part') or '').strip()
                    
                    if not part_val: continue
                    
                    key = f"{sys_val}_{asm_val}_{part_val}".lower()
                    existing[key] = {
                        "system": sys_val,
                        "assembly": asm_val,
                        "part": part_val,
                        "quantity": str(r.get('quantity') or r.get('part_quantity') or '').strip(),
                        "makebuy": str(r.get('make o. buy') or r.get('makebuy') or '').strip().lower()[:1],
                        "comments": str(r.get('comments') or r.get('part_comments') or '').strip(),
                        "id": r.get('DT_RowId')
                    }
            log(f"Found {len(existing)} existing parts on the website.")
        except Exception as e:
            log(f"Could not read existing parts: {e}", "WARN")

        # ── 6. Upload loop ───────────────────────────────────────────────
        from rich.live import Live
        from rich.progress import Progress, SpinnerColumn, BarColumn, TextColumn

        success = 0
        failed = 0
        skipped_dup = 0
        start_time = time.time()

        status_table = Table(title="Live Upload Status", box=None)
        status_table.add_column("Row", justify="right", style="cyan")
        status_table.add_column("Part Name", style="white")
        status_table.add_column("Status", justify="center")
        status_table.add_column("Message", style="dim")

        progress = Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            expand=True
        )
        task_id = progress.add_task("Uploading parts...", total=len(filtered))

        with Live(Panel(Group(status_table, progress), title="BOM Upload Dashboard"), refresh_per_second=4) as live:
            for i, item in enumerate(filtered):
                # Rate Limiting: Burst cooldown
                if i > 0 and i % BURST_LIMIT == 0:
                    log(f"Burst limit reached. Taking a {BURST_COOLDOWN}s break...", "WAIT")
                    status_table.add_row("", "[bold yellow]BURST COOLDOWN[/]", "---", f"Waiting {BURST_COOLDOWN}s")
                    live.update(Panel(Group(status_table, progress), title="BOM Upload Dashboard"))
                    smart_delay(BURST_COOLDOWN)
                
                sys_code = item["system"]
                asm_raw = item["assembly"]
                part_name = item["part"]
                row_num = item["row"]

                # Deduplication check
                dup_key = f"{sys_code}_{asm_raw}_{part_name}".lower()
                
                status_row = [str(row_num), part_name]
                
                is_duplicate = False
                existing_item = existing.get(dup_key)
                
                if existing_item:
                    # Check for "Partial Match" (same key, different data)
                    local_data = {
                        "quantity": str(item["quantity"]),
                        "makebuy": str(item["makebuy"]),
                        "comments": str(item["comments"])
                    }
                    remote_data = {
                        "quantity": str(existing_item["quantity"]),
                        "makebuy": str(existing_item["makebuy"]),
                        "comments": str(existing_item["comments"])
                    }
                    
                    if local_data == remote_data:
                        log(f"Row {row_num}: Exact Duplicate — '{part_name}' already exists", "SKIP")
                        status_row.extend(["[blue]➜ SKIPPED[/]", "Exact match found"])
                        skipped_dup += 1
                        is_duplicate = True
                    else:
                        # Partial Match - Ask User
                        live.stop() # Pause dashboard to ask question
                        from rich.console import Console
                        from rich.panel import Panel
                        Console().print("\n")
                        Console().print(Panel(
                            f"[bold yellow]Partial Match Found for '{part_name}'[/]\n"
                            f"[bold]Excel:[/] Qty={local_data['quantity']}, M/B={local_data['makebuy']}, Note={local_data['comments']}\n"
                            f"[bold]Site: [/] Qty={remote_data['quantity']}, M/B={remote_data['makebuy']}, Note={remote_data['comments']}",
                            title="Conflict Detected"
                        ))
                        
                        action = prompt_ask(
                            questionary.select(
                                "What should I do?",
                                choices=[
                                    questionary.Choice("Skip (Keep site data)", "skip"),
                                    questionary.Choice("Update (Edit existing part)", "update"),
                                    questionary.Choice("Overwrite (Delete & Re-create)", "overwrite")
                                ]
                            )
                        )
                        live.start()
                        
                        if action == "skip":
                            status_row.extend(["[blue]➜ SKIPPED[/]", "User chose to skip"])
                            skipped_dup += 1
                            is_duplicate = True
                        elif action == "update":
                            try:
                                log(f"Row {row_num}: Updating '{part_name}'...", "INFO")
                                row_selector = f"tr#{existing_item['id']}"
                                page.locator(row_selector).click()
                                page.get_by_text("Edit", exact=True).click()
                                page.wait_for_selector(".DTE_Action_Edit")

                                # Only fill fields that differ to save time/requests
                                if local_data["quantity"] != remote_data["quantity"]:
                                    page.locator("#DTE_Field_quantity").fill(local_data["quantity"])
                                if local_data["makebuy"] != remote_data["makebuy"]:
                                    if local_data["makebuy"] == "m":
                                        page.locator("#DTE_Field_makebuy_0").check()
                                    else:
                                        page.locator("#DTE_Field_makebuy_1").check()
                                if local_data["comments"] != remote_data["comments"]:
                                    page.locator("#DTE_Field_comments").fill(local_data["comments"])

                                if DRY_RUN:
                                    log(f"Row {row_num}: [DRY RUN] Would update '{part_name}'", "DRY")
                                    status_row.extend(["[yellow]⚙ DRY RUN[/]", "Simulated update"])
                                    page.keyboard.press("Escape")
                                    is_duplicate = True # skip create logic
                                    success += 1
                                else:
                                    page.get_by_text("Update", exact=True).click()
                                    page.wait_for_selector(".DTE_Action_Edit", state="hidden")
                                    log(f"Row {row_num}: ✓ Updated '{part_name}'", "OK")
                                    status_row.extend(["[blue]✓ UPDATED[/]", "Edited on FSG"])
                                    is_duplicate = True # skip create logic
                                    success += 1
                            except Exception as e:
                                log(f"Row {row_num}: Could not update part: {e}", "WARN")
                                # Fall back to create logic if update fails? 
                                # Better to let is_duplicate=False and try creating a fresh one.
                                is_duplicate = False 
                        elif action == "overwrite":
                            # Delete logic
                            try:
                                log(f"Row {row_num}: Overwriting '{part_name}'...", "INFO")
                                row_selector = f"tr#{existing_item['id']}"
                                page.locator(row_selector).click()
                                page.get_by_text("Delete", exact=True).click()
                                page.locator(".DTE_Action_Remove").get_by_text("Delete").click()
                                page.wait_for_selector(".DTE_Action_Remove", state="hidden")
                                log(f"Row {row_num}: Deleted existing part", "OK")
                            except Exception as e:
                                log(f"Row {row_num}: Could not delete existing part: {e}", "WARN")
                                
                if not is_duplicate:
                    # (Rest of the upload/create logic)
                    attempts = 0
                    uploaded_successfully = False
                    
                    while attempts <= MAX_RETRIES and not uploaded_successfully:
                        try:
                            # Open "New" form
                            page.get_by_text("New", exact=True).click()
                            page.wait_for_selector(".DTE_Action_Create")

                            # System
                            sys_label = SYSTEM_MAP.get(sys_code, sys_code)
                            if not fuzzy_select(page, "#DTE_Field_system", sys_label):
                                raise RuntimeError(f"System '{sys_label}' not found")
                            
                            page.locator("#DTE_Field_system").dispatch_event("change")
                            page.wait_for_timeout(1000)

                            # Assembly
                            if not fuzzy_select(page, "#DTE_Field_assembly", asm_raw):
                                raise RuntimeError(f"Assembly '{asm_raw}' not found")
                            
                            page.locator("#DTE_Field_assembly").dispatch_event("change")

                            # Part name
                            page.locator("#DTE_Field_part").fill(part_name)

                            # Make / Buy
                            if item["makebuy"] == "m":
                                page.locator("#DTE_Field_makebuy_0").check()
                            else:
                                page.locator("#DTE_Field_makebuy_1").check()

                            # Optional fields
                            if item["comments"]:
                                page.locator("#DTE_Field_comments").fill(item["comments"])
                            if item["quantity"]:
                                page.locator("#DTE_Field_quantity").fill(item["quantity"])

                            # Submit (or simulate when DRY_RUN enabled)
                            if DRY_RUN:
                                log(f"Row {row_num}: [DRY RUN] Would create '{part_name}'", "DRY")
                                status_row.extend(["[yellow]⚙ DRY RUN[/]", "Simulated creation"])
                                existing[dup_key] = item # Update local mirror
                                success += 1
                                uploaded_successfully = True
                                page.keyboard.press("Escape")
                            else:
                                page.get_by_text("Create", exact=True).click()
                                page.wait_for_selector(".DTE_Action_Create", state="hidden", timeout=10000)

                                log(f"Row {row_num}: ✓ '{part_name}'", "OK")
                                status_row.extend(["[green]✓ SUCCESS[/]", "Created on FSG"])
                                existing[dup_key] = item # Update local mirror
                                success += 1
                                uploaded_successfully = True

                        except Exception as e:
                            attempts += 1
                            err_msg = str(e).split("\n")[0]
                            log(f"Row {row_num}: Attempt {attempts} failed — {err_msg}", "WARN")
                            
                            try:
                                page.keyboard.press("Escape")
                                page.wait_for_timeout(500)
                            except Exception: pass

                            if attempts > MAX_RETRIES:
                                # Fail-Safe Prompt
                                live.stop()
                                from rich.console import Console
                                from rich.panel import Panel
                                Console().print("\n")
                                Console().print(Panel(
                                    f"[bold red]Failed to process row {row_num} after {attempts} attempts.[/]\n"
                                    f"[bold]Part:[/] {part_name}\n"
                                    f"[bold]Error:[/] {err_msg}",
                                    title="Critical Error"
                                ))
                                
                                final_action = prompt_ask(
                                    questionary.select(
                                        "How to proceed?",
                                        choices=[
                                            questionary.Choice("Retry one more time", "retry"),
                                            questionary.Choice("Skip this part", "skip"),
                                            questionary.Choice("Abort session", "abort")
                                        ]
                                    )
                                )
                                
                                if final_action == "retry":
                                    attempts = MAX_RETRIES # Allow one more loop
                                    live.start()
                                elif final_action == "skip":
                                    status_row.extend(["[red]✗ FAILED[/]", "Skipped after errors"])
                                    failed += 1
                                    live.start()
                                    break
                                elif final_action == "abort":
                                    log("Session aborted by user after errors.", "ERROR")
                                    live.stop()
                                    sys.exit(1)
                            else:
                                # Auto-retry delay
                                smart_delay(5 * attempts)

                # Excel Audit Update
                try:
                    status_text = "SKIPPED" if is_duplicate else ("DONE" if uploaded_successfully else "FAILED")
                    # Automation notes from status_row[3]
                    note_text = status_row[3] if len(status_row) > 3 else ""
                    
                    sheet.cell(row=row_num, column=status_col, value=status_text)
                    sheet.cell(row=row_num, column=notes_col, value=note_text)
                    wb.save(filepath)
                except Exception as e:
                    log(f"Row {row_num}: Could not update Excel status: {e}", "WARN")

                # Update dashboard
                if len(status_table.rows) >= 8:
                    status_table.rows.pop(0)
                status_table.add_row(*status_row)
                
                progress.update(task_id, advance=1)
                
                # Render combined display
                from rich.console import Group
                live.update(Panel(Group(status_table, progress), title="BOM Upload Dashboard"))
                
                # Base delay after each item
                smart_delay(BASE_DELAY)

        # ── 7. Summary ───────────────────────────────────────────────────
        elapsed = round(time.time() - start_time, 1)
        log("─" * 60)
        log(f"Done in {elapsed}s — "
            f"{success} uploaded / {skipped_dup} duplicates / {failed} failed")
        log("─" * 60)

        input("\nPress ENTER to close the browser...")
        browser.close()


if __name__ == "__main__":
    main()
