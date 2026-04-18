import pytest
import openpyxl
from fsg_bom.excel import ExcelProcessor

@pytest.fixture
def sample_excel(tmp_path):
    path = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["System", "Assembly", "Part", "Part_Quantity", "Make o. Buy", "Part_Comments"])
    ws.append(["BR", "Calipers", "Caliper Front", "2", "m", "Notes"])
    ws.append(["DT", "Gearbox", "Gear 1", "1", "b", ""])
    ws.append(["BR", "Calipers", "Already Done", "1", "m", ""])
    ws.cell(row=4, column=1).fill = openpyxl.styles.PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    wb.save(str(path))
    return str(path)

def test_filter_rows(sample_excel):
    processor = ExcelProcessor("BOMs")
    rows, stats = processor.process_file(sample_excel, run_system="ALL")
    
    assert len(rows) == 2
    assert rows[0]["system"] == "BR"
    assert rows[0]["part"] == "Caliper Front"
    assert rows[1]["system"] == "DT"

def test_filter_rows_by_system(sample_excel):
    processor = ExcelProcessor("BOMs")
    rows, stats = processor.process_file(sample_excel, run_system="BR")
    assert len(rows) == 1
    assert rows[0]["system"] == "BR"
